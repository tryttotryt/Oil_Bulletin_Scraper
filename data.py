import requests
import os
import pandas as pd
import openpyxl
import warnings
import datetime

# openpyxl will throw a warning with those spreadsheets, we need to ignore it
warnings.simplefilter("ignore")


# Date and bulletin number from which we start collecting data. need to be manually collected from
# https://ec.europa.eu/energy/observatory/reports/List-of-WOB.pdf
date = datetime.date(2015, 1, 12)
start_bulletin = 1735

# choose your desired country's rows in the spreadsheet.
COUNTRY_CODE = "PL"

# Bulletin upto which we collect data.
end_bulletin = 2131

# Keep track of whether a path was already selected
settled_path = False

def get_country_data(spreadsheet, result_path):
    # Get data from the weekly spreadshit
    raw_workbook = openpyxl.load_workbook(filename=spreadsheet)
    sheet = raw_workbook.active

    # Get data from chosen country
    country_rows = []
    for i, row in enumerate(sheet):
        # If the row is current coutry's, record it
        if sheet[i+1][2].value == COUNTRY_CODE:
            country_rows.append(row)

    # Create or open a spreadsheet to store results
    if not os.path.exists(result_path):
        # Create the spreadsheet if it doesn't exist
        result_workbook = create_spreadsheet(country_rows)
    else:
        # Load the workbook if the file exists
        result_workbook = openpyxl.load_workbook(result_path)

    for i, row in enumerate(country_rows):
        # Open the apprioporate sheet for the fuel
        result_worksheet = result_workbook.get_sheet_by_name(country_rows[i][3].value)
        # Add data from each cell in each row of the selected country
        result_worksheet.append(cell.value for cell in row)

    result_workbook.save(result_path)

def create_spreadsheet(country_rows):
    # Create a workbook
    result_workbook = openpyxl.Workbook()
    del result_workbook["Sheet"]
    # Create a sheet for each fuel type
    for row in country_rows:
        # Get the name of the fuel type to create a sheet of that type
        fuel_type = row[3].value
        worksheet = result_workbook.create_sheet(fuel_type)
        # Add a header to each sheet
        worksheet["A1"] = "Prices in force on"
        worksheet["B1"] = "Country Name"
        worksheet["C1"] = "Country EU Code"
        worksheet["D1"] = "Product Name"
        worksheet["E1"] = "Currency Code"
        worksheet["F1"] = "Prices Unit"
        worksheet["G1"] = "Euro exchange rate"
        worksheet["H1"] = "Weekly price with taxes"
        worksheet["I1"] = "Weekly price without taxes"
    return result_workbook

def get_spreadsheet(link, bulletin):
    # Get the spreadsheet from the european website
    raw_data = requests.get(link)
    # Sometimes there are gaps in weeks which would result in an empty file
    if not raw_data:
        return None
    # Bulletins older than no. 1918 are in the old excel format and need to be converted
    if int(bulletin) > 1917:
        spreadsheet = "downloads/" + bulletin + ".xlsx"
        # Check if the file doesn't exist already
        if os.path.isfile(spreadsheet):
            # Open the spreadsheet and check the first cell (to ensure it's not empty)
            workbook = openpyxl.load_workbook(filename=spreadsheet)
            sheet = workbook.active
            if sheet["A1"].value == "Prices in force on":
                print("It's already here")
                return spreadsheet
        xls = False
    else:
        spreadsheet = "downloads/" + bulletin + ".xls"
        # Check if an .xlsx version of that file doesn't already exist
        if os.path.isfile(spreadsheet + "x"):
            # Open the spreadsheet and check the first cell (to ensure it's not empty)
            workbook = openpyxl.load_workbook(filename=spreadsheet + "x")
            sheet = workbook.active
            if sheet["A1"].value == "Prices in force on":
                print("It's already here")
                return spreadsheet + "x"
        # Remember that it is an old file in need of conversion
        xls = True
    # Save it to a file
    with open(spreadsheet, "wb") as output:
        output.write(raw_data.content)
    # If the file is in the xls format unsupported by openpyxl, convert it
    if xls:
        spreadsheet = xls_to_xlsx(spreadsheet)
    return spreadsheet

def xls_to_xlsx(spreadsheet):
    # Get the xls file to pandas and let it do the conversion
    df = pd.read_excel(spreadsheet, header=None)
    # Store the .xls file name to delete it
    xls_spreadsheet = spreadsheet
    # Save the file as .xlsx
    spreadsheet += "x"
    df.to_excel(spreadsheet, index=False, header=False)
    # Delete the .xls file
    os.remove(xls_spreadsheet)
    return spreadsheet

def get_link(date, bulletin):
    # Construct the link from which the spreadsheat will be downloaded
    start = "http://ec.europa.eu/energy/observatory/reports/"
    if int(bulletin) > 1917:
        link = start + date + "_raw_data_" + bulletin + ".xlsx"
    else:
        link = start + date + "_raw_data_" + bulletin + ".xls"
    return link

def get_path_name():
    # Get the path name for a result
    result_num = 0
    while True:
        # Check if that path name is already taken
        if result_num != 0 :
            num = " (" + str(result_num) + ")"
            result_path = "results/results" + COUNTRY_CODE + num + ".xlsx"
            if not os.path.exists(result_path):
                return result_path
        else: 
            result_path = "results/results" + COUNTRY_CODE + ".xlsx"
            if not os.path.exists(result_path):
                return result_path
        result_num += 1

def main(start_bulletin, end_bulletin, date):

    # Get the path to store the result
    result_path = get_path_name()

    # Start from the first bulletin and end at the last one
    while start_bulletin < end_bulletin + 1:

        # Prepare the variables to be made into a link 
        date_string = date.strftime('%Y_%m_%d')
        bulletin = str(start_bulletin)

        # Get the link for the week's bulletin
        link = get_link(date_string, bulletin)
        print(link)
        # Get the whole spreadsheet from that week
        spreadsheet = get_spreadsheet(link, bulletin)
        # If there weren't any data for that week, go one week ahead
        if not spreadsheet:
            date = date + datetime.timedelta(weeks=1)
            continue
        # Extract desired country's data
        get_country_data(spreadsheet, result_path)

        # Get the next week's date for the next bulletin
        date = date + datetime.timedelta(weeks=1)
        start_bulletin += 1





if __name__ == "__main__":
    """ This is executed when run from the command line """
    main(start_bulletin, end_bulletin, date)







